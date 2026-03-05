import os
import re
import sys
import json
import time
import queue
import threading
import traceback
import csv
import glob
import subprocess
from dataclasses import dataclass
from datetime import datetime
from urllib.parse import urlparse

from PySide6 import QtCore, QtGui, QtWidgets

# optional screen-region capture deps
try:
    import mss  # type: ignore
    from PIL import Image  # type: ignore
except Exception:
    mss = None
    Image = None

# optional Excel import
try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None

# optional browser automation
try:
    from playwright.sync_api import sync_playwright  # type: ignore
except Exception:
    sync_playwright = None


APP_NAME = "SCREENSHOT SCRAPER BATCH RUNNER"
APP_VERSION = "v01"
CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.json")


#helpers
def now_ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def now_human() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def norm_domain(netloc: str) -> str:
    netloc = (netloc or "").strip().lower()
    if netloc.startswith("www."):
        netloc = netloc[4:]
    if ":" in netloc:
        netloc = netloc.split(":", 1)[0]
    return netloc


def clean_url(u: str) -> str:
    u = (u or "").strip()
    if not u:
        return ""
    bad = {"n/a", "#n/a", "na", "none", "null"}
    if u.strip().lower() in bad:
        return ""
    if u.lower().startswith(("http://", "https://")):
        return u
    if re.match(r"^[a-z0-9\-\.]+\.[a-z]{2,}(/|$)", u.strip().lower()):
        return "https://" + u
    return ""


def slugify(s: str, max_len: int = 80) -> str:
    s = (s or "").strip()
    s = re.sub(r"^https?://", "", s, flags=re.I)
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_")
    return s[:max_len] if len(s) > max_len else s


def safe_code(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^a-zA-Z0-9\-\._=]+", "_", s)
    return s[:80] if len(s) > 80 else s


def default_config() -> dict:
    return {
        "output_dir": os.path.join(os.path.dirname(__file__), "screenshots"),
        "filename_template": "{n:04}__{code}__{domain}__{url}__{ts}.png",
        "viewport_w": 1600,
        "viewport_h": 900,
        "full_page": True,
        "capture_mode": "viewport",  # viewport or region
        "region": None,  # saved capture box
        "warmup_per_domain": True,
        "warmup_wait_ms": 2000,
        "after_open_wait_ms": 2000,
        "wait_until": "load",  # page wait target
        "disable_http2": True,
        "disable_quic": True,
        "browser_channel": "chromium",  # selected browser flavor
        "browser_executable": "",  # optional explicit browser path
        "use_persistent_profile": True,
        "profile_dir": os.path.join(os.path.dirname(__file__), "profile"),
        "browser_control_mode": "launch",  # direct launch or CDP attach
        "cdp_host": "127.0.0.1",
        "cdp_port": 9222,
        "cdp_launch_close_existing": False,
    }


def load_config() -> dict:
    cfg = default_config()
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                disk = json.load(f)
            if isinstance(disk, dict):
                cfg.update(disk)
        except Exception:
            pass

    # make sure the output folder exists
    outdir = cfg.get("output_dir") or os.path.join(os.path.dirname(__file__), "screenshots")
    try:
        os.makedirs(outdir, exist_ok=True)
    except Exception:
        outdir = os.path.join(os.path.dirname(__file__), "screenshots")
        os.makedirs(outdir, exist_ok=True)
    cfg["output_dir"] = outdir

    prof = cfg.get("profile_dir") or os.path.join(os.path.dirname(__file__), "profile")
    cfg["profile_dir"] = prof
    return cfg


def save_config(cfg: dict) -> None:
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2)
    except Exception:
        pass


def apply_template(tpl: str, n: int, code: str, url: str) -> str:
    domain = norm_domain(urlparse(url).netloc)
    url_slug = slugify(url, 80)
    ts = now_ts()
    mapping = {
        "n": n,
        "i": n,  # kept for older templates
        "code": safe_code(code),
        "domain": domain,
        "url": url_slug,
        "ts": ts,
        "date": ts,
    }
    try:
        return tpl.format(**mapping)
    except Exception:
        return f"{n:04}__{mapping['code']}__{mapping['domain']}__{mapping['url']}__{ts}.png"


def template_glob(tpl: str, n: int, code: str, url: str) -> str:
    """Build a filename pattern for earlier captures of the same job."""
    domain = norm_domain(urlparse(url).netloc)
    url_slug = slugify(url, 80)
    mapping = {
        "n": n,
        "i": n,  # kept for older templates
        "code": safe_code(code),
        "domain": domain,
        "url": url_slug,
        "ts": "__TS__",
        "date": "__DATE__",
    }
    try:
        name = tpl.format(**mapping)
    except Exception:
        name = f"{n:04}__{mapping['code']}__{mapping['domain']}__{mapping['url']}__{mapping['ts']}.png"

    name = name.replace("__TS__", "*").replace("__DATE__", "*")
    if not re.search(r"\.(png|jpg|jpeg|webp)$", name, re.I):
        name = name + ".*"
    return name


def is_captured(outdir: str, tpl: str, j: "Job") -> bool:
    if not outdir or not os.path.isdir(outdir):
        return False
    pat = template_glob(tpl or "", j.n, j.code, j.url)
    full = os.path.join(outdir, pat)
    return len(glob.glob(full)) > 0


def detect_browser_executable(channel: str, configured_path: str = "") -> str:
    """Best-effort detection for browser executable paths.
    Returns configured_path if it exists; otherwise attempts common install locations.
    """
    channel = (channel or "").strip().lower()
    configured_path = (configured_path or "").strip().strip('"')
    if configured_path and os.path.isfile(configured_path):
        return configured_path

    candidates: list[str] = []

    if sys.platform.startswith("win"):
        local = os.environ.get("LOCALAPPDATA", "")
        prog = os.environ.get("PROGRAMFILES", r"C:\Program Files")
        progx = os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)")

        if channel == "brave":
            candidates += [
                os.path.join(prog, "BraveSoftware", "Brave-Browser", "Application", "brave.exe"),
                os.path.join(progx, "BraveSoftware", "Brave-Browser", "Application", "brave.exe"),
                os.path.join(local, "BraveSoftware", "Brave-Browser", "Application", "brave.exe"),
            ]
        if channel in ("duckduckgo", "ddg"):
            candidates += [
                os.path.join(local, "Programs", "DuckDuckGo", "DuckDuckGo.exe"),
                os.path.join(prog, "DuckDuckGo", "DuckDuckGo.exe"),
                os.path.join(progx, "DuckDuckGo", "DuckDuckGo.exe"),
            ]

    elif sys.platform == "darwin":
        if channel == "brave":
            candidates += ["/Applications/Brave Browser.app/Contents/MacOS/Brave Browser"]
        if channel in ("duckduckgo", "ddg"):
            candidates += ["/Applications/DuckDuckGo.app/Contents/MacOS/DuckDuckGo"]

    else:
        if channel == "brave":
            candidates += ["/usr/bin/brave-browser", "/usr/bin/brave"]
        if channel in ("duckduckgo", "ddg"):
            candidates += ["/usr/bin/duckduckgo", "/usr/bin/duckduckgo-browser"]

    for p in candidates:
        try:
            if p and os.path.isfile(p):
                return p
        except Exception:
            continue
    return ""


#data model
@dataclass
class Job:
    n: int
    code: str
    url: str
    domain: str
    status: str = "queued"
    note: str = ""
    captured: bool = False


#region overlay
class SelectionOverlay(QtWidgets.QWidget):
    rectSelected = QtCore.Signal(QtCore.QRect)

    def __init__(self, screen: QtGui.QScreen):
        super().__init__()
        self.setWindowFlags(QtCore.Qt.WindowType.FramelessWindowHint | QtCore.Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(QtCore.Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self._screen = screen
        self._origin = None
        self._rect = QtCore.QRect()

    def paintEvent(self, e):
        p = QtGui.QPainter(self)
        p.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
        p.fillRect(self.rect(), QtGui.QColor(0, 0, 0, 60))
        if not self._rect.isNull():
            pen = QtGui.QPen(QtGui.QColor(0, 200, 255, 255), 2)
            p.setPen(pen)
            p.drawRect(self._rect)

    def mousePressEvent(self, e):
        self._origin = e.position().toPoint()
        self._rect = QtCore.QRect(self._origin, QtCore.QSize())
        self.update()

    def mouseMoveEvent(self, e):
        if self._origin is None:
            return
        pos = e.position().toPoint()
        self._rect = QtCore.QRect(self._origin, pos).normalized()
        self.update()

    def mouseReleaseEvent(self, e):
        if self._origin is None:
            return
        pos = e.position().toPoint()
        self._rect = QtCore.QRect(self._origin, pos).normalized()
        self.update()
        self.rectSelected.emit(self._rect)
        self.close()

    def keyPressEvent(self, e):
        if e.key() == QtCore.Qt.Key.Key_Escape:
            self.close()


#browser worker
class Browser(QtCore.QObject):
    log = QtCore.Signal(str)
    ready = QtCore.Signal(bool)
    task_done = QtCore.Signal(str, bool, object, str)

    def __init__(self, cfg: dict):
        super().__init__()
        self.cfg = cfg
        self._pw = None
        self._browser = None
        self._context = None
        self._page = None
        self._thread = None
        self._tasks = queue.Queue()
        self._stop_evt = threading.Event()
        self._started = False
        self._warm_domains = set()

    def start(self):
        if self._started:
            return
        if sync_playwright is None:
            self.log.emit("ERROR: playwright not installed. Install requirements then run: playwright install chromium")
            self.ready.emit(False)
            return

        # guard against a broken Browser class patch
        assert hasattr(self, "_launch"), "Browser is missing _launch (bad patching?)"

        def worker():
            try:
                self._pw = sync_playwright().start()
                assert hasattr(self, "_launch"), "Browser is missing _launch (bad patching?)"
                self._launch()
                self._started = True
                self.ready.emit(True)

                while not self._stop_evt.is_set():
                    try:
                        task = self._tasks.get(timeout=0.25)
                    except Exception:
                        continue
                    if not task:
                        continue
                    if task.get("kind") == "stop":
                        break

                    tid = task.get("id", "")
                    kind = task.get("kind", "")
                    try:
                        if kind == "relaunch":
                            self._relaunch()
                            self.task_done.emit(tid, True, None, "")
                        elif kind == "goto":
                            self._handle_goto(tid, task)
                        elif kind == "screenshot_viewport":
                            self._handle_shot_viewport(tid, task)
                    except Exception as e:
                        self.task_done.emit(tid, False, None, str(e))

            except Exception as e:
                self.log.emit("Browser worker crash: " + str(e))
                self.log.emit(traceback.format_exc())
                self.ready.emit(False)
            finally:
                try:
                    if self._context:
                        self._context.close()
                    elif self._browser:
                        self._browser.close()
                except Exception:
                    pass
                try:
                    if self._pw:
                        self._pw.stop()
                except Exception:
                    pass
                self._browser = self._context = self._page = self._pw = None
                self._started = False
                self.log.emit("Browser stopped.")

        self._thread = threading.Thread(target=worker, daemon=True)
        self._thread.start()

    def stop(self):
        self._stop_evt.set()
        self._tasks.put({"kind": "stop"})
        try:
            if self._thread:
                self._thread.join(timeout=2)
        except Exception:
            pass

    def submit(self, task: dict) -> str:
        tid = task.get("id") or str(datetime.now().timestamp())
        task["id"] = tid
        self._tasks.put(task)
        return tid

    def _launch(self):
        w = int(self.cfg.get("viewport_w", 1600))
        h = int(self.cfg.get("viewport_h", 900))

        mode_ctl = (self.cfg.get("browser_control_mode", "launch") or "launch").lower().strip()
        if mode_ctl not in ("launch", "cdp"):
            mode_ctl = "launch"

        # in CDP mode we attach to an existing browser session
        if mode_ctl == "cdp":
            host = (self.cfg.get("cdp_host", "127.0.0.1") or "127.0.0.1").strip()
            port = int(self.cfg.get("cdp_port", 9222) or 9222)
            endpoint = f"http://{host}:{port}"
            try:
                self._browser = self._pw.chromium.connect_over_cdp(endpoint)
            except Exception as e:
                raise RuntimeError(
                    f"Failed to connect to existing browser via CDP at {endpoint}. "
                    f"Start Chromium/Brave/Chrome with --remote-debugging-port={port} first. "
                    f"Error: {e}"
                )

            # use the default persistent context when it exists
            ctx = None
            try:
                if getattr(self._browser, "contexts", None) and self._browser.contexts:
                    ctx = self._browser.contexts[0]
            except Exception:
                ctx = None

            if ctx is None:
                ctx = self._browser.new_context(
                    viewport={"width": w, "height": h},
                    device_scale_factor=1,
                )

            self._context = ctx
            try:
                pages = list(getattr(self._context, "pages", []) or [])
            except Exception:
                pages = []
            self._page = pages[0] if pages else self._context.new_page()

            self._warm_domains = set()
            self.log.emit(
                f"Connected via CDP to existing browser at {endpoint}. "
                f"Viewport={w}x{h}. Solve any Cloudflare challenge manually in that window, then capture."
            )
            return

        # regular Playwright launch path
        launch_args = [
            f"--window-size={w+80},{h+140}",
            "--force-device-scale-factor=1",
            "--high-dpi-support=1",
        ]
        if bool(self.cfg.get("disable_http2", True)):
            launch_args.append("--disable-http2")
        if bool(self.cfg.get("disable_quic", True)):
            launch_args.append("--disable-quic")

        channel = (self.cfg.get("browser_channel", "chromium") or "chromium").lower()
        launch_kwargs = dict(headless=False, args=launch_args)

        if channel in ("msedge", "chrome"):
            launch_kwargs["channel"] = channel
        elif channel in ("brave", "duckduckgo"):
            exe = detect_browser_executable(channel, self.cfg.get("browser_executable", ""))
            if exe:
                launch_kwargs["executable_path"] = exe
            else:
                self.log.emit(
                    f"WARNING: Could not auto-detect {channel} executable. "
                    "Set Browser EXE in the UI or switch to chromium."
                )
                channel = "chromium"

        use_persistent = bool(self.cfg.get("use_persistent_profile", True))

        profile_dir = (self.cfg.get("profile_dir") or "").strip()

        if use_persistent and profile_dir:
            os.makedirs(profile_dir, exist_ok=True)
            self._context = self._pw.chromium.launch_persistent_context(
                user_data_dir=profile_dir,
                **launch_kwargs,
                viewport={"width": w, "height": h},
                device_scale_factor=1,
            )
            try:
                self._browser = self._context.browser
            except Exception:
                self._browser = None
            self._page = self._context.pages[0] if getattr(self._context, "pages", None) and self._context.pages else self._context.new_page()
            mode = "persistent"
        else:
            self._browser = self._pw.chromium.launch(**launch_kwargs)
            self._context = self._browser.new_context(
                viewport={"width": w, "height": h},
                device_scale_factor=1,
            )
            self._page = self._context.new_page()
            mode = "ephemeral"

        self._warm_domains = set()
        self.log.emit(
            f"Browser launched (headful, {mode}). Channel={channel} Viewport={w}x{h} "
            f"Flags: http2={'off' if self.cfg.get('disable_http2', True) else 'on'} "
            f"quic={'off' if self.cfg.get('disable_quic', True) else 'on'} "
            + (f"Profile={profile_dir}" if (mode == "persistent") else "")
        )

    def _relaunch(self):
        try:
            # with CDP this usually just disconnects the session
            if self._context:
                try:
                    self._context.close()
                except Exception:
                    pass
            if self._browser:
                try:
                    self._browser.close()
                except Exception:
                    pass
        except Exception:
            pass
        self._browser = self._context = self._page = None
        self._launch()

    def _handle_goto(self, tid: str, task: dict):
        url = task["url"]
        warmup = bool(task.get("warmup", False))
        warm_ms = int(task.get("warmup_wait_ms", 0))
        after_ms = int(task.get("after_open_wait_ms", 0))
        wait_until = task.get("wait_until", "load")
        timeout_ms = int(task.get("timeout_ms", 60000))

        dom = norm_domain(urlparse(url).netloc)
        t0 = time.time()

        try:
            if warmup and dom and dom not in self._warm_domains:
                home = f"https://{dom}/"
                self.log.emit(f"Warm-up {dom}: {home}")
                self._page.goto(home, wait_until="domcontentloaded", timeout=timeout_ms)
                if warm_ms > 0:
                    time.sleep(warm_ms / 1000)
                self._warm_domains.add(dom)

            self.log.emit(f"Open: {url}")
            self._page.goto(url, wait_until=wait_until, timeout=timeout_ms)
            if after_ms > 0:
                time.sleep(after_ms / 1000)

            dt = time.time() - t0
            self.task_done.emit(tid, True, {"duration_s": dt}, "")
        except Exception as e:
            msg = str(e)
            if "ERR_HTTP2_PROTOCOL_ERROR" in msg or "HTTP2_PROTOCOL_ERROR" in msg:
                self.log.emit(f"HTTP/2 protocol error detected. Relaunch + retry for {dom}...")
                self._relaunch()
                try:
                    self._page.goto(url, wait_until=wait_until, timeout=timeout_ms)
                    if after_ms > 0:
                        time.sleep(after_ms / 1000)
                    dt = time.time() - t0
                    self.task_done.emit(tid, True, {"duration_s": dt, "retried": True}, "")
                    return
                except Exception as e2:
                    self.task_done.emit(tid, False, None, str(e2))
                    return
            self.task_done.emit(tid, False, None, msg)

    def _handle_shot_viewport(self, tid: str, task: dict):
        path = task["path"]
        full_page = bool(task.get("full_page", True))
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            self._page.screenshot(path=path, full_page=full_page)
            self.task_done.emit(tid, True, {"path": path}, "")
        except Exception as e:
            self.task_done.emit(tid, False, None, str(e))


def browser_unit_self_check() -> None:
    """Fail early if the Browser class is missing _launch."""
    import inspect
    methods = sorted([k for k, v in Browser.__dict__.items() if inspect.isfunction(v) and not k.startswith("__")])
    print(f"[{APP_NAME}] Browser methods: " + ", ".join(methods))
    if "_launch" not in Browser.__dict__ or not callable(Browser.__dict__.get("_launch")):
        print(f"[{APP_NAME}] FATAL: Browser._launch missing or not callable. Exiting.")
        raise SystemExit(2)


#main window
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.cfg = load_config()
        self.jobs: list[Job] = []
        self.current_idx = 0
        self._overlay = None
        self._pending_kind = ""

        self.setWindowTitle(f"{APP_NAME} {APP_VERSION}")
        self.resize(1200, 800)

        self.browser = Browser(self.cfg)
        self.browser.log.connect(self.add_log)
        self.browser.ready.connect(self.on_browser_ready)
        self.browser.task_done.connect(self.on_task_done)

        self._build_ui()
        self._apply_cfg_to_ui()

        self.browser.start()
        self.add_log(f"Ready. {APP_NAME} {APP_VERSION}. {now_human()}")

    #main ui
    def _build_ui(self):
        cw = QtWidgets.QWidget()
        self.setCentralWidget(cw)
        root = QtWidgets.QVBoxLayout(cw)

        top = QtWidgets.QHBoxLayout()
        self.btn_load_excel = QtWidgets.QPushButton("Load Excel…")
        self.btn_load_excel.clicked.connect(self.load_excel)
        top.addWidget(self.btn_load_excel)

        self.btn_paste = QtWidgets.QPushButton("Paste Rows…")
        self.btn_paste.clicked.connect(self.paste_rows)
        top.addWidget(self.btn_paste)

        self.lbl_out = QtWidgets.QLabel("")
        top.addWidget(self.lbl_out, 1)

        self.btn_out = QtWidgets.QPushButton("Output Folder…")
        self.btn_out.clicked.connect(self.pick_output_dir)
        top.addWidget(self.btn_out)

        self.btn_open_folder = QtWidgets.QPushButton("Open Folder")
        self.btn_open_folder.clicked.connect(self.open_output_dir)
        top.addWidget(self.btn_open_folder)

        root.addLayout(top)

        mid = QtWidgets.QHBoxLayout()

        left = QtWidgets.QVBoxLayout()
        self.lst = QtWidgets.QListWidget()
        self.lst.currentRowChanged.connect(self.on_row_changed)
        left.addWidget(self.lst, 1)

        prog = QtWidgets.QHBoxLayout()
        self.lbl_prog = QtWidgets.QLabel("0 / 0")
        prog.addWidget(self.lbl_prog)
        self.lbl_last = QtWidgets.QLabel("Last open: –")
        prog.addWidget(self.lbl_last, 1)
        left.addLayout(prog)
        mid.addLayout(left, 2)

        panel = QtWidgets.QVBoxLayout()

        #timing
        grp_t = QtWidgets.QGroupBox("Timing")
        gtl = QtWidgets.QGridLayout(grp_t)
        self.chk_warm = QtWidgets.QCheckBox("Warm-up per domain")
        self.chk_warm.setToolTip("Before first deep link on a domain, visit its homepage once.")
        gtl.addWidget(self.chk_warm, 0, 0, 1, 2)

        self.chk_dedupe = QtWidgets.QCheckBox("Remove duplicates (code+link)")
        self.chk_dedupe.setToolTip("When loading data, drop exact duplicate (code + URL) pairs.")
        self.chk_dedupe.setChecked(True)
        gtl.addWidget(self.chk_dedupe, 0, 2, 1, 2)

        self.spin_warm = QtWidgets.QSpinBox(); self.spin_warm.setRange(0, 60000); self.spin_warm.setSuffix(" ms")
        self.spin_after = QtWidgets.QSpinBox(); self.spin_after.setRange(0, 60000); self.spin_after.setSuffix(" ms")
        gtl.addWidget(QtWidgets.QLabel("Warm-up wait"), 1, 0); gtl.addWidget(self.spin_warm, 1, 1)
        gtl.addWidget(QtWidgets.QLabel("After-open wait"), 2, 0); gtl.addWidget(self.spin_after, 2, 1)

        self.cmb_wait = QtWidgets.QComboBox()
        self.cmb_wait.addItems(["domcontentloaded", "load", "networkidle"])
        self.cmb_wait.setToolTip("What counts as 'page loaded' before the after-open wait starts. 'load' is usually best.")
        gtl.addWidget(QtWidgets.QLabel("Wait until"), 3, 0); gtl.addWidget(self.cmb_wait, 3, 1)

        panel.addWidget(grp_t)

        #capture settings
        grp_c = QtWidgets.QGroupBox("Capture")
        gcl = QtWidgets.QGridLayout(grp_c)
        self.cmb_mode = QtWidgets.QComboBox()
        self.cmb_mode.addItems(["viewport", "region"])
        self.cmb_mode.setToolTip("Viewport: browser screenshot. Region: capture a screen rectangle you pick.")
        gcl.addWidget(QtWidgets.QLabel("Mode"), 0, 0); gcl.addWidget(self.cmb_mode, 0, 1)

        self.chk_full = QtWidgets.QCheckBox("Full page")
        self.chk_full.setToolTip("Capture full scrollable page (recommended).")
        gcl.addWidget(self.chk_full, 1, 0, 1, 2)

        self.spin_w = QtWidgets.QSpinBox(); self.spin_w.setRange(600, 4000)
        self.spin_h = QtWidgets.QSpinBox(); self.spin_h.setRange(400, 3000)
        gcl.addWidget(QtWidgets.QLabel("Viewport W"), 2, 0); gcl.addWidget(self.spin_w, 2, 1)
        gcl.addWidget(QtWidgets.QLabel("Viewport H"), 3, 0); gcl.addWidget(self.spin_h, 3, 1)

        self.btn_apply_viewport = QtWidgets.QPushButton("Apply Viewport (relaunch)")
        self.btn_apply_viewport.setToolTip("Changes screenshot size. Relaunch browser so the window matches the viewport.")
        self.btn_apply_viewport.clicked.connect(self.apply_viewport)
        gcl.addWidget(self.btn_apply_viewport, 4, 0, 1, 2)

        self.btn_pick_region = QtWidgets.QPushButton("Pick Region…")
        self.btn_pick_region.setToolTip("Drag a rectangle on your screen. Used only when Mode=region.")
        self.btn_pick_region.clicked.connect(self.pick_region)
        gcl.addWidget(self.btn_pick_region, 5, 0, 1, 2)

        self.lbl_region = QtWidgets.QLabel("Region: (not set)")
        gcl.addWidget(self.lbl_region, 6, 0, 1, 2)

        panel.addWidget(grp_c)

        #file naming
        grp_f = QtWidgets.QGroupBox("File naming")
        gfl = QtWidgets.QVBoxLayout(grp_f)
        self.txt_tpl = QtWidgets.QLineEdit()
        self.txt_tpl.setToolTip("Fields: {n} index, {code}, {domain}, {url} (slug), {ts}/{date}. Default: {n:04}__{code}__{domain}__{url}__{ts}.png")
        gfl.addWidget(self.txt_tpl)
        self.lbl_preview = QtWidgets.QLabel("Preview: ")
        gfl.addWidget(self.lbl_preview)
        panel.addWidget(grp_f)

        #browser settings
        grp_b = QtWidgets.QGroupBox("Browser")
        gbl = QtWidgets.QGridLayout(grp_b)

        self.cmb_channel = QtWidgets.QComboBox()
        self.cmb_channel.addItems(["chromium", "msedge", "chrome", "brave", "duckduckgo"])
        self.cmb_channel.setToolTip("Chromium default. Edge/Chrome use Playwright channels. Brave/DDG require an EXE path (auto-detect best-effort).")
        gbl.addWidget(QtWidgets.QLabel("Channel"), 0, 0); gbl.addWidget(self.cmb_channel, 0, 1)

        self.cmb_control = QtWidgets.QComboBox()
        self.cmb_control.addItems(["launch", "cdp"])
        self.cmb_control.setToolTip("launch = Playwright launches a browser. cdp = attach to an already-running Chromium browser started with --remote-debugging-port.")
        gbl.addWidget(QtWidgets.QLabel("Control"), 1, 0)
        gbl.addWidget(self.cmb_control, 1, 1, 1, 2)

        self.txt_cdp_host = QtWidgets.QLineEdit()
        self.txt_cdp_host.setPlaceholderText("127.0.0.1")
        self.spin_cdp_port = QtWidgets.QSpinBox()
        self.spin_cdp_port.setRange(1, 65535)
        self.spin_cdp_port.setValue(9222)

        row_cdp = QtWidgets.QHBoxLayout()
        row_cdp.addWidget(QtWidgets.QLabel("Host"))
        row_cdp.addWidget(self.txt_cdp_host, 1)
        row_cdp.addWidget(QtWidgets.QLabel("Port"))
        row_cdp.addWidget(self.spin_cdp_port)
        w_cdp = QtWidgets.QWidget(); w_cdp.setLayout(row_cdp)
        gbl.addWidget(QtWidgets.QLabel("CDP"), 2, 0)
        gbl.addWidget(w_cdp, 2, 1, 1, 2)

        self.btn_launch_cdp = QtWidgets.QPushButton("Launch CDP Browser")
        self.btn_launch_cdp.setToolTip("Launch the selected browser with --remote-debugging-port so the app can attach in Control=cdp mode.")
        self.btn_launch_cdp.clicked.connect(self.launch_cdp_browser)
        gbl.addWidget(self.btn_launch_cdp, 3, 0, 1, 3)


        self.txt_browser_exe = QtWidgets.QLineEdit()
        self.txt_browser_exe.setPlaceholderText("Optional: browser executable path (for Brave / DuckDuckGo / custom)")
        btn_pick_exe = QtWidgets.QPushButton("Pick…")
        btn_pick_exe.clicked.connect(self.pick_browser_exe)
        gbl.addWidget(QtWidgets.QLabel("Browser EXE"), 4, 0)
        gbl.addWidget(self.txt_browser_exe, 4, 1)
        gbl.addWidget(btn_pick_exe, 4, 2)

        self.chk_profile = QtWidgets.QCheckBox("Use persistent profile (user-data-dir)")
        self.chk_profile.setToolTip("Keeps cookies/login between runs using Playwright persistent context.")
        gbl.addWidget(self.chk_profile, 5, 0, 1, 3)

        self.txt_profile_dir = QtWidgets.QLineEdit()
        self.txt_profile_dir.setPlaceholderText("Profile folder (user-data-dir)")
        btn_pick_profile = QtWidgets.QPushButton("Pick…")
        btn_pick_profile.clicked.connect(self.pick_profile_dir)
        gbl.addWidget(QtWidgets.QLabel("Profile Dir"), 6, 0)
        gbl.addWidget(self.txt_profile_dir, 6, 1)
        gbl.addWidget(btn_pick_profile, 6, 2)

        self.chk_http2 = QtWidgets.QCheckBox("Disable HTTP/2")
        self.chk_http2.setToolTip("Launch flag to avoid some HTTP/2 protocol errors (e.g., CDW). Requires relaunch.")
        self.chk_quic = QtWidgets.QCheckBox("Disable QUIC")
        self.chk_quic.setToolTip("Launch flag to disable QUIC. Requires relaunch.")
        gbl.addWidget(self.chk_http2, 7, 0, 1, 3)
        gbl.addWidget(self.chk_quic, 8, 0, 1, 3)

        self.btn_relaunch = QtWidgets.QPushButton("Relaunch Browser")
        self.btn_relaunch.clicked.connect(self.relaunch_browser)
        gbl.addWidget(self.btn_relaunch, 9, 0, 1, 3)

        panel.addWidget(grp_b)

        self.cmb_control.currentTextChanged.connect(self.on_cfg_changed)
        self.txt_cdp_host.textChanged.connect(self.on_cfg_changed)
        self.spin_cdp_port.valueChanged.connect(self.on_cfg_changed)
        self.cmb_control.currentTextChanged.connect(self._update_browser_controls)
        self._update_browser_controls()

        #actions
        btns = QtWidgets.QHBoxLayout()
        self.btn_open = QtWidgets.QPushButton("Open (F7)")
        self.btn_cap = QtWidgets.QPushButton("Capture (F8)")
        self.btn_skip = QtWidgets.QPushButton("Skip (F9)")
        self.btn_open.clicked.connect(self.open_current)
        self.btn_cap.clicked.connect(self.capture_current)
        self.btn_skip.clicked.connect(self.skip_current)
        btns.addWidget(self.btn_open)
        btns.addWidget(self.btn_cap)
        btns.addWidget(self.btn_skip)

        self.btn_export = QtWidgets.QPushButton("Export Report")
        self.btn_export.setToolTip("Export the rows list to a CSV capture report (includes captured flag).")
        self.btn_export.clicked.connect(self.export_report)
        btns.addWidget(self.btn_export)
        panel.addLayout(btns)

        panel.addStretch(1)
        mid.addLayout(panel, 1)

        root.addLayout(mid, 1)

        self.txt_log = QtWidgets.QPlainTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setMaximumBlockCount(3000)
        root.addWidget(self.txt_log, 1)

        #hotkeys
        QtGui.QShortcut(QtGui.QKeySequence("F7"), self, activated=self.open_current)
        QtGui.QShortcut(QtGui.QKeySequence("F8"), self, activated=self.capture_current)
        QtGui.QShortcut(QtGui.QKeySequence("F9"), self, activated=self.skip_current)

        QtGui.QShortcut(QtGui.QKeySequence("1"), self, activated=self.open_current)
        QtGui.QShortcut(QtGui.QKeySequence("2"), self, activated=self.capture_current)
        QtGui.QShortcut(QtGui.QKeySequence("3"), self, activated=self.skip_current)
        QtGui.QShortcut(QtGui.QKeySequence("Num1"), self, activated=self.open_current)
        QtGui.QShortcut(QtGui.QKeySequence("Num2"), self, activated=self.capture_current)
        QtGui.QShortcut(QtGui.QKeySequence("Num3"), self, activated=self.skip_current)

        #save config when fields change
        self.txt_tpl.textChanged.connect(self.on_cfg_changed)
        self.txt_browser_exe.textChanged.connect(self.on_cfg_changed)
        self.txt_profile_dir.textChanged.connect(self.on_cfg_changed)
        for w in [
            self.chk_warm, self.spin_warm, self.spin_after, self.cmb_wait,
            self.cmb_mode, self.chk_full, self.spin_w, self.spin_h,
            self.chk_http2, self.chk_quic, self.cmb_channel, self.chk_profile
        ]:
            if isinstance(w, QtWidgets.QAbstractButton):
                w.toggled.connect(self.on_cfg_changed)
            elif isinstance(w, QtWidgets.QComboBox):
                w.currentIndexChanged.connect(self.on_cfg_changed)
            else:
                w.valueChanged.connect(self.on_cfg_changed)

    def _apply_cfg_to_ui(self):
        self.lbl_out.setText(f"Output: {self.cfg.get('output_dir')}")
        self.txt_tpl.setText(self.cfg.get("filename_template", ""))
        self.spin_w.setValue(int(self.cfg.get("viewport_w", 1600)))
        self.spin_h.setValue(int(self.cfg.get("viewport_h", 900)))
        self.chk_full.setChecked(bool(self.cfg.get("full_page", True)))
        self.cmb_mode.setCurrentText(self.cfg.get("capture_mode", "viewport"))
        self.chk_warm.setChecked(bool(self.cfg.get("warmup_per_domain", True)))
        self.spin_warm.setValue(int(self.cfg.get("warmup_wait_ms", 2000)))
        self.spin_after.setValue(int(self.cfg.get("after_open_wait_ms", 2000)))
        self.cmb_wait.setCurrentText(self.cfg.get("wait_until", "load"))
        self.chk_http2.setChecked(bool(self.cfg.get("disable_http2", True)))
        self.chk_quic.setChecked(bool(self.cfg.get("disable_quic", True)))
        self.cmb_channel.setCurrentText(self.cfg.get("browser_channel", "chromium"))
        self.cmb_control.setCurrentText(self.cfg.get("browser_control_mode", "launch"))
        self.txt_cdp_host.setText(self.cfg.get("cdp_host", "127.0.0.1"))
        self.spin_cdp_port.setValue(int(self.cfg.get("cdp_port", 9222)))
        self.txt_browser_exe.setText(self.cfg.get("browser_executable", ""))
        self.chk_profile.setChecked(bool(self.cfg.get("use_persistent_profile", True)))
        self.txt_profile_dir.setText(self.cfg.get("profile_dir", ""))
        self._update_region_label()
        self._update_preview()

    def _update_region_label(self):
        r = self.cfg.get("region")
        if r and isinstance(r, dict):
            self.lbl_region.setText(f"Region: x={r['x']} y={r['y']} w={r['w']} h={r['h']}")
        else:
            self.lbl_region.setText("Region: (not set)")

    def _update_preview(self):
        if not self.jobs:
            sample_n, sample_code, sample_url = 1, "CODE", "https://example.com/product/123"
        else:
            j = self.jobs[self.current_idx] if 0 <= self.current_idx < len(self.jobs) else self.jobs[0]
            sample_n, sample_code, sample_url = j.n, j.code, j.url
        tpl = self.txt_tpl.text().strip()
        self.lbl_preview.setText("Preview: " + apply_template(tpl, sample_n, sample_code, sample_url))

    #logging and browser callbacks
    def add_log(self, msg: str):
        self.txt_log.appendPlainText(f"[{now_human()}] {msg}")

    def on_browser_ready(self, ok: bool):
        if ok:
            self.add_log("Browser is ready. Use Open/Capture to work through the queue.")
        else:
            self.add_log("Browser failed to start. See log above.")

    def on_task_done(self, tid: str, ok: bool, result: object, err: str):
        if self._pending_kind == "goto":
            self.btn_open.setEnabled(True)
            self.btn_cap.setEnabled(True)
            if ok:
                dur = result.get("duration_s") if isinstance(result, dict) else None
                if dur is not None:
                    self.lbl_last.setText(f"Last open: {dur:.1f}s")
                self._update_preview()
            else:
                self.add_log("Navigation error: " + err)
        elif self._pending_kind == "shot":
            if ok:
                path = result.get("path") if isinstance(result, dict) else ""
                self.add_log(f"Saved: {path}")
                self.mark_status("saved", "")
                self.advance()
            else:
                self.add_log("Screenshot error: " + err)
                self.mark_status("error", err)

    #job list and progress
    def refresh_list(self):
        self.lst.blockSignals(True)
        self.lst.clear()

        tpl = self.cfg.get("filename_template", "")
        outdir = self.cfg.get("output_dir", "")

        for j in self.jobs:
            try:
                j.captured = is_captured(outdir, tpl, j)
            except Exception:
                j.captured = False
            cap = "Captured" if j.captured else "—"
            self.lst.addItem(f"{cap} | {j.n:04} | {j.code} | {j.domain} | {j.status}")

        self.lst.blockSignals(False)

    def update_progress(self):
        total = len(self.jobs)
        self.lbl_prog.setText(f"{self.current_idx + 1 if total else 0} / {total}")

    def on_row_changed(self, row: int):
        if 0 <= row < len(self.jobs):
            self.current_idx = row
            self.update_progress()
            self._update_preview()

    def mark_status(self, status: str, note: str):
        if 0 <= self.current_idx < len(self.jobs):
            self.jobs[self.current_idx].status = status
            self.jobs[self.current_idx].note = note
            self.refresh_list()

    def advance(self):
        if self.current_idx < len(self.jobs) - 1:
            self.current_idx += 1
            self.lst.setCurrentRow(self.current_idx)
        self.update_progress()

    #config sync
    def on_cfg_changed(self):
        self.cfg["filename_template"] = self.txt_tpl.text().strip()
        self.cfg["viewport_w"] = int(self.spin_w.value())
        self.cfg["viewport_h"] = int(self.spin_h.value())
        self.cfg["full_page"] = bool(self.chk_full.isChecked())
        self.cfg["capture_mode"] = self.cmb_mode.currentText()
        self.cfg["warmup_per_domain"] = bool(self.chk_warm.isChecked())
        self.cfg["warmup_wait_ms"] = int(self.spin_warm.value())
        self.cfg["after_open_wait_ms"] = int(self.spin_after.value())
        self.cfg["wait_until"] = self.cmb_wait.currentText()
        self.cfg["disable_http2"] = bool(self.chk_http2.isChecked())
        self.cfg["disable_quic"] = bool(self.chk_quic.isChecked())
        self.cfg["browser_channel"] = self.cmb_channel.currentText()
        self.cfg["browser_control_mode"] = self.cmb_control.currentText()
        self.cfg["cdp_host"] = self.txt_cdp_host.text().strip() or "127.0.0.1"
        self.cfg["cdp_port"] = int(self.spin_cdp_port.value())
        self.cfg["browser_executable"] = self.txt_browser_exe.text().strip()
        self.cfg["use_persistent_profile"] = bool(self.chk_profile.isChecked())
        self.cfg["profile_dir"] = self.txt_profile_dir.text().strip()
        save_config(self.cfg)
        self._update_preview()
        self.refresh_list()


    def _update_browser_controls(self):
        mode_ctl = (self.cmb_control.currentText() or "launch").lower().strip()
        is_cdp = (mode_ctl == "cdp")
        self.txt_cdp_host.setEnabled(is_cdp)
        self.spin_cdp_port.setEnabled(is_cdp)
        self.btn_launch_cdp.setEnabled(is_cdp)
        # CDP relaunch here really means reconnect
        self.btn_relaunch.setText("Reconnect Browser" if is_cdp else "Relaunch Browser")

    def launch_cdp_browser(self):
        # start a Chromium-based browser with remote debugging enabled
        exe = self.txt_browser_exe.text().strip()
        if not exe:
            # fall back to the selected browser channel
            channel = (self.cmb_channel.currentText() or "chromium").lower().strip()
            exe = detect_browser_executable(channel, "")
        if not exe or not os.path.exists(exe):
            self.add_log("CDP launch: Browser EXE not set or not found. Use 'Pick…' to select brave.exe/chrome.exe/etc.")
            return

        host = (self.txt_cdp_host.text().strip() or "127.0.0.1").strip()
        port = int(self.spin_cdp_port.value())
        profile_dir = (self.txt_profile_dir.text().strip() or "").strip()
        if not profile_dir:
            profile_dir = os.path.join(os.path.dirname(__file__), "profile")
            self.txt_profile_dir.setText(profile_dir)
            self.cfg["profile_dir"] = profile_dir
            save_config(self.cfg)

        os.makedirs(profile_dir, exist_ok=True)

        w = int(self.spin_w.value())
        h = int(self.spin_h.value())

        args = [
            exe,
            f"--remote-debugging-port={port}",
            f"--user-data-dir={profile_dir}",
            f"--window-size={w+80},{h+140}",
        ]

        try:
            subprocess.Popen(args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, close_fds=(os.name != "nt"))
            self.add_log(f"Launched CDP browser: {' '.join(args)}")
            self.add_log(f"Now set Control=cdp and click Reconnect Browser to attach to http://{host}:{port}")
        except Exception as e:
            self.add_log(f"CDP launch failed: {e}")


    #pickers
    def pick_browser_exe(self):
        filt = "Executable (*.exe);;All Files (*)" if sys.platform.startswith("win") else "All Files (*)"
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select browser executable", "", filt)
        if not path:
            return
        self.txt_browser_exe.setText(path)
        self.cfg["browser_executable"] = path
        save_config(self.cfg)
        self.add_log(f"Browser EXE set: {path}")

    def pick_profile_dir(self):
        start = self.cfg.get("profile_dir") or os.path.dirname(__file__)
        d = QtWidgets.QFileDialog.getExistingDirectory(self, "Select profile folder (user-data-dir)", start)
        if not d:
            return
        self.txt_profile_dir.setText(d)
        self.cfg["profile_dir"] = d
        save_config(self.cfg)
        self.add_log(f"Profile dir set: {d}")

    def pick_output_dir(self):
        d = QtWidgets.QFileDialog.getExistingDirectory(self, "Select output folder", self.cfg.get("output_dir"))
        if not d:
            return
        self.cfg["output_dir"] = d
        os.makedirs(d, exist_ok=True)
        save_config(self.cfg)
        self.lbl_out.setText(f"Output: {d}")
        self.add_log(f"Output folder set: {d}")

    def open_output_dir(self):
        path = self.cfg.get("output_dir")
        QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(path))

    #browser controls
    def apply_viewport(self):
        self.on_cfg_changed()
        self.relaunch_browser()

    def relaunch_browser(self):
        self.browser.cfg = self.cfg
        self._pending_kind = "relaunch"
        self.browser.submit({"kind": "relaunch"})
        self.add_log("Relaunch requested (applies channel/HTTP2/QUIC/viewport/profile).")

    #region capture
    def pick_region(self):
        if mss is None or Image is None:
            QtWidgets.QMessageBox.warning(self, "Region capture not available", "Region capture requires mss + Pillow.")
            return
        self.cmb_mode.setCurrentText("region")
        screen = QtGui.QGuiApplication.primaryScreen()
        if not screen:
            self.add_log("ERROR: no screen found.")
            return
        self.add_log("Pick Region: click and drag to select capture area.")
        self._overlay = SelectionOverlay(screen)
        self._overlay.rectSelected.connect(self.on_region_selected)
        self._overlay.showFullScreen()
        self._overlay.activateWindow()
        self._overlay.raise_()

    def on_region_selected(self, r: QtCore.QRect):
        self.cfg["region"] = {"x": int(r.left()), "y": int(r.top()), "w": int(r.width()), "h": int(r.height())}
        save_config(self.cfg)
        self._overlay = None
        self._update_region_label()
        self.add_log(f"Selected region: x={r.left()} y={r.top()} w={r.width()} h={r.height()}")

    #workflow
    def current_job(self):
        if 0 <= self.current_idx < len(self.jobs):
            return self.jobs[self.current_idx]
        return None

    def open_current(self):
        j = self.current_job()
        if not j:
            return
        self.on_cfg_changed()
        self.browser.cfg = self.cfg
        self._pending_kind = "goto"
        self.btn_open.setEnabled(False)
        self.btn_cap.setEnabled(False)
        self.browser.submit({
            "kind": "goto",
            "url": j.url,
            "warmup": bool(self.cfg.get("warmup_per_domain", True)),
            "warmup_wait_ms": int(self.cfg.get("warmup_wait_ms", 0)),
            "after_open_wait_ms": int(self.cfg.get("after_open_wait_ms", 0)),
            "wait_until": self.cfg.get("wait_until", "load"),
            "timeout_ms": 60000,
        })
        j.status = "opened"
        self.refresh_list()

    def capture_current(self):
        j = self.current_job()
        if not j:
            return
        self.on_cfg_changed()
        outdir = self.cfg.get("output_dir")
        os.makedirs(outdir, exist_ok=True)
        filename = apply_template(self.cfg.get("filename_template", ""), j.n, j.code, j.url)
        outpath = os.path.join(outdir, filename)

        if self.cfg.get("capture_mode") == "region":
            r = self.cfg.get("region")
            if not (r and isinstance(r, dict) and r.get("w") and r.get("h")):
                QtWidgets.QMessageBox.information(self, "Region not set", "Pick a region first.")
                return
            try:
                with mss.mss() as sct:
                    mon = {"left": int(r["x"]), "top": int(r["y"]), "width": int(r["w"]), "height": int(r["h"])}
                    shot = sct.grab(mon)
                    img = Image.frombytes("RGB", shot.size, shot.rgb)
                    img.save(outpath)
                self.add_log(f"Saved (region): {outpath}")
                self.mark_status("saved", "")
                self.advance()
            except Exception as e:
                self.add_log("Region screenshot error: " + str(e))
                self.mark_status("error", str(e))
            return

        self.browser.cfg = self.cfg
        self._pending_kind = "shot"
        self.browser.submit({
            "kind": "screenshot_viewport",
            "path": outpath,
            "full_page": bool(self.cfg.get("full_page", True)),
        })

    def skip_current(self):
        j = self.current_job()
        if j:
            j.status = "skipped"
            self.refresh_list()
        self.advance()

    #export
    def export_report(self):
        if not self.jobs:
            QtWidgets.QMessageBox.information(self, "No data", "There are no rows to export.")
            return

        try:
            self.refresh_list()
        except Exception:
            pass

        outdir = self.cfg.get("output_dir") or os.path.join(os.getcwd(), "screenshots")
        os.makedirs(outdir, exist_ok=True)

        ts = now_ts()
        default_path = os.path.join(outdir, f"capture_report__{ts}.csv")
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save Capture Report", default_path, "CSV Files (*.csv)")
        if not path:
            return

        tpl = self.cfg.get("filename_template", "")
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["captured", "n", "code", "domain", "url", "status", "note", "expected_glob"])
                for j in self.jobs:
                    cap = is_captured(outdir, tpl, j)
                    expected = template_glob(tpl, j.n, j.code, j.url)
                    w.writerow(["yes" if cap else "no", j.n, j.code, j.domain, j.url, j.status, j.note, expected])
            self.add_log(f"Exported capture report: {path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export failed", str(e))
            self.add_log("Export failed: " + str(e))

    #import
    def load_excel(self):
        if openpyxl is None:
            QtWidgets.QMessageBox.warning(self, "openpyxl not installed", "Excel import requires openpyxl.")
            return
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select Excel workbook", "", "Excel (*.xlsx)")
        if not path:
            return
        wb = openpyxl.load_workbook(path, data_only=True)
        pairs = []
        for ws in wb.worksheets:
            try:
                headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
            except Exception:
                continue
            if not headers:
                continue
            code_idx = 0
            for i, h in enumerate(headers):
                hl = h.lower()
                if any(k in hl for k in ["part", "product", "code", "sku", "pn", "part number"]):
                    code_idx = i
                    break
            link_cols = [i for i, h in enumerate(headers) if "link" in h.lower() or "url" in h.lower()]
            if not link_cols:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                code = safe_code(str(row[code_idx]) if code_idx < len(row) and row[code_idx] is not None else "")
                if not code:
                    continue
                for li in link_cols:
                    if li >= len(row):
                        continue
                    u = row[li]
                    if not isinstance(u, str):
                        continue
                    cu = clean_url(u)
                    if cu:
                        pairs.append((code, cu))
        self._load_jobs_from_pairs(pairs, source=os.path.basename(path))

    def paste_rows(self):
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Paste rows")
        dlg.resize(800, 500)
        v = QtWidgets.QVBoxLayout(dlg)
        txt = QtWidgets.QPlainTextEdit()
        txt.setPlaceholderText("Paste tab-separated rows: code<TAB>url\nYou can also paste one URL per line.")
        v.addWidget(txt, 1)
        btns = QtWidgets.QHBoxLayout()
        ok = QtWidgets.QPushButton("Load")
        cancel = QtWidgets.QPushButton("Cancel")
        btns.addStretch(1)
        btns.addWidget(ok)
        btns.addWidget(cancel)
        v.addLayout(btns)
        cancel.clicked.connect(dlg.reject)

        def do_load():
            raw = txt.toPlainText().strip()
            if not raw:
                dlg.reject()
                return
            pairs = []
            for line in raw.splitlines():
                line = line.strip()
                if not line:
                    continue
                if "\t" in line:
                    a, b = line.split("\t", 1)
                    code = safe_code(a)
                    url = clean_url(b)
                    if code and url:
                        pairs.append((code, url))
                else:
                    url = clean_url(line)
                    if url:
                        pairs.append(("", url))
            dlg.accept()
            self._load_jobs_from_pairs(pairs, source="paste")

        ok.clicked.connect(do_load)
        dlg.exec()

    def _load_jobs_from_pairs(self, pairs, source=""):
        self.jobs = []
        self.current_idx = 0

        if self.chk_dedupe.isChecked():
            seen = set()
            deduped = []
            for c, u in pairs:
                key = (str(c).strip(), str(u).strip())
                if key in seen:
                    continue
                seen.add(key)
                deduped.append((c, u))
            pairs = deduped

        n = 0
        for code, url in pairs:
            url = clean_url(url)
            if not url:
                continue
            n += 1
            dom = norm_domain(urlparse(url).netloc)
            self.jobs.append(Job(n=n, code=code or f"ITEM{n}", url=url, domain=dom))

        self.refresh_list()
        if self.jobs:
            self.lst.setCurrentRow(0)
        self.update_progress()
        self.add_log(f"Loaded {len(self.jobs)} jobs from {source}.")

    # backup numeric hotkeys for systems that miss the normal bindings
    def keyPressEvent(self, event: QtGui.QKeyEvent):
        k = event.key()
        mods = event.modifiers()
        if k == QtCore.Qt.Key.Key_1:
            self.open_current(); return
        if k == QtCore.Qt.Key.Key_2:
            self.capture_current(); return
        if k == QtCore.Qt.Key.Key_3:
            self.skip_current(); return
        if (mods & QtCore.Qt.KeyboardModifier.KeypadModifier) and k == QtCore.Qt.Key.Key_1:
            self.open_current(); return
        if (mods & QtCore.Qt.KeyboardModifier.KeypadModifier) and k == QtCore.Qt.Key.Key_2:
            self.capture_current(); return
        if (mods & QtCore.Qt.KeyboardModifier.KeypadModifier) and k == QtCore.Qt.Key.Key_3:
            self.skip_current(); return
        super().keyPressEvent(event)

    def closeEvent(self, e: QtGui.QCloseEvent):
        try:
            self.on_cfg_changed()
        except Exception:
            pass
        try:
            self.browser.stop()
        except Exception:
            pass
        return super().closeEvent(e)


def main():
    browser_unit_self_check()
    app = QtWidgets.QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
